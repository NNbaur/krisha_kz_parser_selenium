import bs4
from typing import Type
from bs4 import BeautifulSoup
import glob
from openpyxl import Workbook
from openpyxl.styles import Alignment


def get_characteristics(soup: Type[bs4.BeautifulSoup]) -> list:
    list_of_chars = []
    # Characteristics of flats from html
    chars = [
        'flat.building',
        'flat.floor',
        'live.square',
        'map.complex',
        'house.year'
    ]
    # Characteristics are optional, so some of them may be not in ad
    for char in chars:
        result = None
        char_block = soup.find('div', {
            'class': 'offer__info-item',
            'data-name': char
        })
        # If char. is noticed, get text and append to result list
        if char_block:
            result = char_block.find(
                'div',
                class_="offer__advert-short-info"
            ).get_text(strip=True)
        list_of_chars.append(result)
    return list_of_chars


def get_main_info(soup: Type[bs4.BeautifulSoup]) -> list:
    main_info = []
    classes = [
        "offer__location offer__advert-short-info",
        "offer__advert-title",
        "offer__price",
        "text"
    ]
    for i in classes:
        # Location in ad is contained in the following class in span tag
        if i.startswith("offer__location"):
            info = soup.find(('p', 'div'), class_=i).span.get_text(strip=True)
        else:
            info = soup.find(('p', 'div'), class_=i).get_text(strip=True)
        main_info.append(info)
    return main_info


def get_data_from_offers() -> list:
    full_data = []
    flat_pages = glob.glob("../pages/flats/*.html")

    for flat in flat_pages:
        with open(flat, 'r', encoding='UTF-8') as f:
            html = f.read()
            soup = BeautifulSoup(html, 'html.parser')
            main_info = get_main_info(soup)
            chars = get_characteristics(soup)
            full_data.append(main_info + chars)
    return full_data


def save_to_excel():
    print('Please wait. Saving data to excel file')
    data = get_data_from_offers()
    wb = Workbook()
    ws = wb.active

    # Make column titles
    col_titles = {
        "A1": 'Местоположение', "B1": 'Заголовок', "C1": 'Цена',
        "D1": 'Описание', "E1": 'Тип здания', "F1": 'Этаж',
        "G1": 'Площадь', "H1": 'Жилой комплекс', "I1": 'Год постройки'
    }
    for k, v in col_titles.items():
        ws[k] = v

    # Fill data in worksheet
    for d in data:
        ws.append(d)

    cell_range2 = ws['A1':f'I{len(data)+1}']
    # Formatting cells
    for cells in cell_range2:
        for cell in cells:
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )

    col_width = {
        "A": 21, "B": 42, "C": 14,
        "D": 85, "E": 14, "F": 8,
        "G": 16, "H": 18, "I": 14
    }
    # Set width of columns
    for k, v in col_width.items():
        ws.column_dimensions[k].width = v

    wb.save('../flats.xlsx')
    print('Excel file is created')
